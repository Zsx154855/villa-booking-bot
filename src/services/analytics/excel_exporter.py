#!/usr/bin/env python3
"""
Taimili Villa Booking System - Excel Exporter
Excel报表导出器

使用openpyxl生成Excel格式的数据报表
"""

import os
import sys
import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Any

# 尝试导入openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 添加项目根目录到路径
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, PROJECT_ROOT)

# 配置日志
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel报表导出器"""
    
    # 样式定义
    HEADER_FONT = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    CELL_FONT = Font(name='微软雅黑', size=10)
    TITLE_FONT = Font(name='微软雅黑', size=16, bold=True)
    SUBTITLE_FONT = Font(name='微软雅黑', size=11, italic=True)
    
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
    CELL_ALIGN = Alignment(horizontal='left', vertical='center')
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 交替行颜色
    ODD_FILL = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    EVEN_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    def __init__(self):
        """初始化Excel导出器"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl未安装，请运行: pip install openpyxl")
    
    def _create_workbook(self) -> Workbook:
        """创建工作簿"""
        return Workbook()
    
    def _apply_header_style(self, cell):
        """应用表头样式"""
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = self.HEADER_ALIGN
        cell.border = self.THIN_BORDER
    
    def _apply_cell_style(self, cell, is_odd: bool = True):
        """应用单元格样式"""
        cell.font = self.CELL_FONT
        cell.fill = self.ODD_FILL if is_odd else self.EVEN_FILL
        cell.alignment = self.CELL_ALIGN
        cell.border = self.THIN_BORDER
    
    def _apply_title_style(self, cell):
        """应用标题样式"""
        cell.font = self.TITLE_FONT
        cell.alignment = self.HEADER_ALIGN
    
    def _auto_adjust_column_width(self, ws, min_width: int = 10, max_width: int = 50):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _write_header_row(self, ws, headers: List[str], row: int = 1):
        """写入表头行"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_header_style(cell)
    
    def _write_data_row(self, ws, data: List[Any], row: int, formats: List[str] = None):
        """写入数据行"""
        formats = formats or []
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            self._apply_cell_style(cell, is_odd=(row % 2 == 1))
            
            # 应用格式化
            fmt = formats[col - 1] if col <= len(formats) else None
            if fmt and value is not None:
                if fmt == 'money':
                    cell.number_format = '¥#,##0.00'
                elif fmt == 'percent':
                    cell.number_format = '0.00%'
                elif fmt == 'int':
                    cell.number_format = '#,##0'
    
    # ============ 日报表导出 ============
    
    def export_daily_report(self, report: Dict, output_path: str = None) -> str:
        """
        导出日报表到Excel
        
        Args:
            report: 日报表数据
            output_path: 输出文件路径
        
        Returns:
            输出文件路径
        """
        wb = self._create_workbook()
        ws = wb.active
        ws.title = '日报表'
        
        # 标题
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"日报表 - {report['date_display']}"
        self._apply_title_style(title_cell)
        
        # 概览数据
        ws['A3'] = '指标'
        ws['B3'] = '数值'
        self._apply_header_style(ws['A3'])
        self._apply_header_style(ws['B3'])
        
        overview_data = [
            ('总预订数', report['total_bookings']),
            ('总收入', report['total_revenue']),
        ]
        
        for i, (label, value) in enumerate(overview_data, 4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
            self._apply_cell_style(ws.cell(row=i, column=1), is_odd=(i % 2 == 0))
            self._apply_cell_style(ws.cell(row=i, column=2), is_odd=(i % 2 == 0))
        
        # 地区分布表
        start_row = 7
        ws.cell(row=start_row, column=1, value='地区分布').font = Font(bold=True, size=12)
        
        headers = ['地区', '预订数', '收入']
        self._write_header_row(ws, headers, start_row + 1)
        
        for i, region in enumerate(report.get('region_distribution', []), start_row + 2):
            self._write_data_row(ws, [
                region['region'],
                region['bookings'],
                region['revenue']
            ], i, formats=[None, 'int', 'money'])
        
        # 热门别墅表
        if report.get('top_villas'):
            villa_start = start_row + len(report['region_distribution']) + 3
            ws.cell(row=villa_start, column=1, value='热门别墅排行').font = Font(bold=True, size=12)
            
            headers = ['排名', '别墅名称', '地区', '预订数', '收入']
            self._write_header_row(ws, headers, villa_start + 1)
            
            for i, villa in enumerate(report['top_villas'], villa_start + 2):
                self._write_data_row(ws, [
                    villa['rank'],
                    villa['name'],
                    villa['region'],
                    villa['bookings'],
                    villa['revenue']
                ], i, formats=['int', None, None, 'int', 'money'])
        
        # 调整列宽
        self._auto_adjust_column_width(ws)
        
        # 保存文件
        if not output_path:
            output_dir = os.path.join(PROJECT_ROOT, 'reports')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"daily_report_{report['date']}.xlsx")
        
        wb.save(output_path)
        logger.info(f"✅ 日报表已导出: {output_path}")
        
        return output_path
    
    # ============ 周报表导出 ============
    
    def export_weekly_report(self, report: Dict, output_path: str = None) -> str:
        """
        导出周报表到Excel
        
        Args:
            report: 周报表数据
            output_path: 输出文件路径
        
        Returns:
            输出文件路径
        """
        wb = self._create_workbook()
        ws = wb.active
        ws.title = '周报表'
        
        # 标题
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"周报表 - {report['date_range_display']}"
        self._apply_title_style(title_cell)
        
        # 概览
        ws['A3'] = '指标'
        ws['B3'] = '数值'
        self._apply_header_style(ws['A3'])
        self._apply_header_style(ws['B3'])
        
        overview_data = [
            ('总预订数', report['total_bookings']),
            ('总收入', report['total_revenue']),
            ('日均预订', round(report['total_bookings'] / max(len(report['daily_trends']), 1), 2)),
            ('日均收入', round(report['total_revenue'] / max(len(report['daily_trends']), 1), 2)),
        ]
        
        for i, (label, value) in enumerate(overview_data, 4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
            self._apply_cell_style(ws.cell(row=i, column=1), is_odd=(i % 2 == 0))
            self._apply_cell_style(ws.cell(row=i, column=2), is_odd=(i % 2 == 0))
        
        # 每日趋势表
        start_row = 9
        ws.cell(row=start_row, column=1, value='每日趋势').font = Font(bold=True, size=12)
        
        headers = ['日期', '预订数', '收入']
        self._write_header_row(ws, headers, start_row + 1)
        
        for i, day in enumerate(report.get('daily_trends', []), start_row + 2):
            self._write_data_row(ws, [
                day['date'],
                day['bookings'],
                day['revenue']
            ], i, formats=[None, 'int', 'money'])
        
        # 地区分布
        if report.get('region_distribution'):
            region_start = start_row + len(report['daily_trends']) + 3
            ws.cell(row=region_start, column=1, value='地区分布').font = Font(bold=True, size=12)
            
            headers = ['地区', '预订数', '收入', '占比']
            self._write_header_row(ws, headers, region_start + 1)
            
            total_bookings = sum(r['bookings'] for r in report['region_distribution'])
            for i, region in enumerate(report['region_distribution'], region_start + 2):
                pct = region['bookings'] / total_bookings if total_bookings else 0
                self._write_data_row(ws, [
                    region['region'],
                    region['bookings'],
                    region['revenue'],
                    pct
                ], i, formats=[None, 'int', 'money', 'percent'])
        
        # 客户来源
        if report.get('customer_sources'):
            source_start = region_start + len(report['region_distribution']) + 3
            ws.cell(row=source_start, column=1, value='客户来源').font = Font(bold=True, size=12)
            
            headers = ['来源', '订单数', '收入']
            self._write_header_row(ws, headers, source_start + 1)
            
            for i, source in enumerate(report['customer_sources'], source_start + 2):
                self._write_data_row(ws, [
                    source['source'],
                    source['count'],
                    source['revenue']
                ], i, formats=[None, 'int', 'money'])
        
        self._auto_adjust_column_width(ws)
        
        if not output_path:
            output_dir = os.path.join(PROJECT_ROOT, 'reports')
            os.makedirs(output_dir, exist_ok=True)
            filename = f"weekly_report_{report['start_date']}_{report['end_date']}.xlsx"
            output_path = os.path.join(output_dir, filename)
        
        wb.save(output_path)
        logger.info(f"✅ 周报表已导出: {output_path}")
        
        return output_path
    
    # ============ 月报表导出 ============
    
    def export_monthly_report(self, report: Dict, output_path: str = None) -> str:
        """
        导出月报表到Excel
        
        Args:
            report: 月报表数据
            output_path: 输出文件路径
        
        Returns:
            输出文件路径
        """
        wb = self._create_workbook()
        ws = wb.active
        ws.title = '月报表'
        
        # 标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"月报表 - {report['month_display']}"
        self._apply_title_style(title_cell)
        
        # 概览
        ws['A3'] = '指标'
        ws['B3'] = '数值'
        self._apply_header_style(ws['A3'])
        self._apply_header_style(ws['B3'])
        
        overview_data = [
            ('总预订数', report['total_bookings']),
            ('总收入', report['total_revenue']),
            ('平均客单价', report['avg_price']),
            ('日均预订', round(report['total_bookings'] / max(len(report['daily_trends']), 1), 2)),
        ]
        
        for i, (label, value) in enumerate(overview_data, 4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
            self._apply_cell_style(ws.cell(row=i, column=1), is_odd=(i % 2 == 0))
            self._apply_cell_style(ws.cell(row=i, column=2), is_odd=(i % 2 == 0))
        
        # 同比环比分析
        comp = report['comparison']
        start_row = 9
        
        # 环比
        ws.cell(row=start_row, column=1, value='环比变化 (上月)').font = Font(bold=True, size=12, color='2E86AB')
        mom = comp['month_over_month']
        
        headers = ['指标', '上月', '本月', '变化']
        self._write_header_row(ws, headers, start_row + 1)
        
        self._write_data_row(ws, [
            '预订数', mom['prev_bookings'], report['total_bookings'],
            f"{mom['bookings_change']:+.2f}%"
        ], start_row + 2)
        self._write_data_row(ws, [
            '收入', f"¥{mom['prev_revenue']:,.2f}", f"¥{report['total_revenue']:,.2f}",
            f"{mom['revenue_change']:+.2f}%"
        ], start_row + 3)
        
        # 同比
        yoy_start = start_row + 5
        ws.cell(row=yoy_start, column=1, value='同比变化 (去年同月)').font = Font(bold=True, size=12, color='A23B72')
        yoy = comp['year_over_year']
        
        self._write_header_row(ws, headers, yoy_start + 1)
        
        self._write_data_row(ws, [
            '预订数', yoy['yoy_bookings'], report['total_bookings'],
            f"{yoy['bookings_change']:+.2f}%"
        ], yoy_start + 2)
        self._write_data_row(ws, [
            '收入', f"¥{yoy['yoy_revenue']:,.2f}", f"¥{report['total_revenue']:,.2f}",
            f"{yoy['revenue_change']:+.2f}%"
        ], yoy_start + 3)
        
        # 每日趋势表
        trend_start = yoy_start + 5
        ws.cell(row=trend_start, column=1, value='每日趋势').font = Font(bold=True, size=12)
        
        headers = ['日期', '预订数', '收入']
        self._write_header_row(ws, headers, trend_start + 1)
        
        for i, day in enumerate(report.get('daily_trends', []), trend_start + 2):
            self._write_data_row(ws, [
                day['date'],
                day['bookings'],
                day['revenue']
            ], i, formats=[None, 'int', 'money'])
        
        # 地区分布
        if report.get('region_distribution'):
            region_start = trend_start + len(report['daily_trends']) + 3
            ws.cell(row=region_start, column=1, value='地区分布').font = Font(bold=True, size=12)
            
            headers = ['地区', '预订数', '收入', '占比']
            self._write_header_row(ws, headers, region_start + 1)
            
            total_bookings = sum(r['bookings'] for r in report['region_distribution'])
            for i, region in enumerate(report['region_distribution'], region_start + 2):
                pct = region['bookings'] / total_bookings if total_bookings else 0
                self._write_data_row(ws, [
                    region['region'],
                    region['bookings'],
                    region['revenue'],
                    pct
                ], i, formats=[None, 'int', 'money', 'percent'])
        
        self._auto_adjust_column_width(ws)
        
        if not output_path:
            output_dir = os.path.join(PROJECT_ROOT, 'reports')
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"monthly_report_{report['year']}_{report['month']:02d}.xlsx")
        
        wb.save(output_path)
        logger.info(f"✅ 月报表已导出: {output_path}")
        
        return output_path
    
    # ============ 综合报表导出 ============
    
    def export_comprehensive_report(
        self, 
        daily: Dict = None, 
        weekly: Dict = None, 
        monthly: Dict = None,
        output_path: str = None
    ) -> str:
        """
        导出综合报表（包含日、周、月数据）
        
        Args:
            daily: 日报表数据
            weekly: 周报表数据
            monthly: 月报表数据
            output_path: 输出文件路径
        
        Returns:
            输出文件路径
        """
        wb = self._create_workbook()
        
        # 日报表sheet
        if daily:
            ws_daily = wb.active
            ws_daily.title = '日报表'
            self._export_sheet_data(ws_daily, daily, 'daily')
        
        # 周报表sheet
        if weekly:
            ws_weekly = wb.create_sheet('周报表')
            self._export_sheet_data(ws_weekly, weekly, 'weekly')
        
        # 月报表sheet
        if monthly:
            ws_monthly = wb.create_sheet('月报表')
            self._export_sheet_data(ws_monthly, monthly, 'monthly')
        
        if not output_path:
            output_dir = os.path.join(PROJECT_ROOT, 'reports')
            os.makedirs(output_dir, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(output_dir, f"comprehensive_report_{timestamp}.xlsx")
        
        wb.save(output_path)
        logger.info(f"✅ 综合报表已导出: {output_path}")
        
        return output_path
    
    def _export_sheet_data(self, ws, report: Dict, report_type: str):
        """导出sheet数据"""
        # 简单实现 - 写入关键数据
        row = 1
        
        if report_type == 'daily':
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value=f"日报表 - {report.get('date_display', '')}")
            row += 2
            
            ws.cell(row=row, column=1, value='总预订数')
            ws.cell(row=row, column=2, value=report.get('total_bookings', 0))
            row += 1
            ws.cell(row=row, column=1, value='总收入')
            ws.cell(row=row, column=2, value=report.get('total_revenue', 0))
        
        elif report_type == 'weekly':
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value=f"周报表 - {report.get('date_range_display', '')}")
            row += 2
            
            ws.cell(row=row, column=1, value='总预订数')
            ws.cell(row=row, column=2, value=report.get('total_bookings', 0))
            row += 1
            ws.cell(row=row, column=1, value='总收入')
            ws.cell(row=row, column=2, value=report.get('total_revenue', 0))
        
        elif report_type == 'monthly':
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1, value=f"月报表 - {report.get('month_display', '')}")
            row += 2
            
            ws.cell(row=row, column=1, value='总预订数')
            ws.cell(row=row, column=2, value=report.get('total_bookings', 0))
            row += 1
            ws.cell(row=row, column=1, value='总收入')
            ws.cell(row=row, column=2, value=report.get('total_revenue', 0))
        
        self._auto_adjust_column_width(ws)


# ============ 命令行入口 ============
if __name__ == '__main__':
    import argparse
    
    # 导入report_generator
    from report_generator import ReportGenerator
    
    parser = argparse.ArgumentParser(description='导出别墅预订Excel报表')
    parser.add_argument('--type', '-t', choices=['daily', 'weekly', 'monthly', 'all'],
                        default='daily', help='报表类型')
    parser.add_argument('--date', '-d', help='日期 (YYYY-MM-DD)')
    parser.add_argument('--output', '-o', help='输出文件路径')
    args = parser.parse_args()
    
    if not OPENPYXL_AVAILABLE:
        print("❌ 错误: openpyxl未安装，请运行: pip install openpyxl")
        sys.exit(1)
    
    generator = ReportGenerator()
    exporter = ExcelExporter()
    
    if args.date:
        report_date = datetime.strptime(args.date, '%Y-%m-%d').date()
    else:
        report_date = None
    
    output_path = args.output
    
    if args.type == 'daily':
        report = generator.get_daily_report(report_date)
        path = exporter.export_daily_report(report, output_path)
        print(f"✅ 日报表已导出: {path}")
    
    elif args.type == 'weekly':
        report = generator.get_weekly_report(report_date)
        path = exporter.export_weekly_report(report, output_path)
        print(f"✅ 周报表已导出: {path}")
    
    elif args.type == 'monthly':
        year = report_date.year if report_date else None
        month = report_date.month if report_date else None
        report = generator.get_monthly_report(year, month)
        path = exporter.export_monthly_report(report, output_path)
        print(f"✅ 月报表已导出: {path}")
    
    elif args.type == 'all':
        daily = generator.get_daily_report(report_date)
        weekly = generator.get_weekly_report(report_date)
        year = report_date.year if report_date else None
        month = report_date.month if report_date else None
        monthly = generator.get_monthly_report(year, month)
        
        path = exporter.export_comprehensive_report(daily, weekly, monthly, output_path)
        print(f"✅ 综合报表已导出: {path}")
    
    print("\n✨ 完成!")
